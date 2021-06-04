#player positions run from 0 - 39 (need to mod 40 on every roll)
#0 - go, 1 = medeter ave, 2 = comm chest ... 39 = boardwalk
#0 = go, 10 = jail, 20 = free parking, 30 = go to jail
#railroads in the middle of each lane @ 5, 15, 25, 35
#utilities 2 away from jails @ 12 & 28
#tax @ 4 & 38
#chance at 8, 22, 26.
#comm chest @ 2, 17, 33

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import openpyxl
import random

def shuffle(value): #chance and comm chest card only [1,16]

    result = {}

    for i in range(1, 17):
        rand_val = random.randint(1, 16) #16 cards for chance/comm chest
        
        while (value[rand_val] is None): #Ensure no repeats in the result
            rand_val = random.randint(1, 16)
            
        result[i] = value[rand_val]
        value[rand_val] = None
        
    return result

def newPos(curr, card_val): 
    ''' Check based on current position if there is a needed move. 
        Eg: Chance card, comm chest card
    '''

    if (0 <= card_val <= 39): #move to point on board
        curr = card_val
    
    elif (0 > card_val): #back 3 spaces
        curr += card_val

    elif (card_val == 41): #utility
        if (13 <= curr <= 28):
            curr = 28
            
        else:
            curr = 12

    elif (card_val == 42): #railroad
        if (6 <= curr <= 15):
            curr = 15

        elif (16 <= curr <= 25):
            curr = 25

        elif (26 <= curr <= 35):
            curr = 35

        else:
            curr = 5

    return curr % 40
    
def clean(value): #print the dictionary nicely

    length = len(value) - 1
    
    first_key = list(value.keys())[0]
    last_key = list(value.keys())[length]
    
    for i in range(first_key, last_key+1):
        print('{}: {}'.format(i, round(value[i],2)))

    return ''
    
def rollDice(): #individual dice to allow for doubles
    
    return random.randint(1,6)

def community(comm_bank): #increment the deck

    first_value = comm_bank[1]

    for i in range (1, 16):
        comm_bank[i] = comm_bank[i+1]

    comm_bank[16] = first_value

    return comm_bank

def chance(chance_cards): #increment the deck
    
    first_value = chance_cards[1]

    for i in range (1, 16):
        chance_cards[i] = chance_cards[i+1]

    chance_cards[16] = first_value

    return chance_cards

'''-----------------------------------------------------------------------------------'''
#wb = Workbook()
#wb.save(filename = 'Monopoly.xlsx')
wb = load_workbook('Monopoly.xlsx')
sheet = wb.active

''' ------------------------------------------- Formatting starts here ------------------------------------------- '''

head_1 = sheet.cell(row = 1, column = 1)
head_1.value = 'Location'
head_1.font = Font(name = 'Times New Roman', size = 12, bold = True)
head_1.alignment = Alignment(horizontal="center", vertical="center")

head_2 = sheet.cell(row = 1, column = 2)
head_2.value = 'Count'
head_2.font = Font(name = 'Times New Roman', size = 12, bold = True)
head_2.alignment = Alignment(horizontal="center", vertical="center")

head_3 = sheet.cell(row = 1, column = 3)
head_3.value = 'Percentage'
head_3.font = Font(name = 'Times New Roman', size = 12, bold = True)
head_3.alignment = Alignment(horizontal="center", vertical="center")

sheet.column_dimensions['A'].width = 19.5
sheet.column_dimensions['B'].width = 19.5
sheet.column_dimensions['C'].width = 19.5

boardName = ['Go', 'Mediterranean Avenue', 'Community Chest', 'Baltic Avenue', 'Income Tax', 'Reading Railroad', 'Oriental Avenue', 'Chance', 'Vermont Avenue', 'Connecticut Avenue', 'Jail', 'St. Charles Place', 'Electric Company', 'States Avenue', 'Virginia Avenue', 'Pennsylvania Railroad', 'St. James Place', 'Tennessee Avenue', 'New York Avenue', 'Free Parking', 'Kentucky Avenue', 'Indiana Avenue', 'Illinois Avenue', 'B & O Railroad', 'Atlantic Avenue', 'Ventnor Avenue', 'Water Works', 'Marvin Gardens', 'Go To Jail', 'Pacific Avenue', 'North Carolina Avenue', 'Pennsylvania Avenue','Short Line', 'Park Place', 'Luxury Tax', 'Boardwalk']

for i in range(0, len(boardName)):
    sheet.cell(row = i+2, column = 1).value = boardName[i]
    sheet.cell(row = i+2, column = 1).font = Font(name = 'Times New Roman', size = 12)
    sheet.cell(row = i+2, column = 1).alignment = Alignment(horizontal="center", vertical="center")
    
    sheet.cell(row = i+2, column = 2).font = Font(name = 'Times New Roman', size = 12)
    sheet.cell(row = i+2, column = 2).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row = i+2, column = 2).value = None
    
    sheet.cell(row = i+2, column = 3).font = Font(name = 'Times New Roman', size = 12)
    sheet.cell(row = i+2, column = 3).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row = i+2, column = 3).value = None
    
for i in range(2, 38):
    
    if (i == 3 or i == 5):
        sheet.cell(row = i, column = 1).fill = PatternFill(start_color='DEB887', end_color='DEB887',fill_type='solid')

    elif (i == 8 or i == 10 or i == 11):
        sheet.cell(row = i, column = 1).fill = PatternFill(start_color='87CEEB', end_color='87CEEB',fill_type='solid')

    elif (i == 13 or i == 15 or i == 16):
        sheet.cell(row = i, column = 1).fill = PatternFill(start_color='FF1493', end_color='FF1493',fill_type='solid')

    elif (i == 18 or i == 19 or i == 20):
        sheet.cell(row = i, column = 1).fill = PatternFill(start_color='FF8C00', end_color='FF8C00',fill_type='solid')

    elif (i == 22 or i == 23 or i == 24):
        sheet.cell(row = i, column = 1).fill = PatternFill(start_color='FF0000', end_color='FF0000',fill_type='solid')

    elif (i == 26 or i == 27 or i == 29):
        sheet.cell(row = i, column = 1).fill = PatternFill(start_color='FFFF00', end_color='FFFF00',fill_type='solid')

    elif (i == 31 or i == 32 or i == 33):
        sheet.cell(row = i, column = 1).fill = PatternFill(start_color='2E8B57', end_color='2E8B57',fill_type='solid')

    elif (i == 37 or i == 35):
        sheet.cell(row = i, column = 1).fill = PatternFill(start_color='4169E1', end_color='4169E1',fill_type='solid')

    else:
        sheet.cell(row = i, column = 1).fill = PatternFill(fill_type = None)

''' ------------------------------------------- Formatting ends here -------------------------------------------'''

num_players = 4

comm_chest = {1: 0, 2: 10, 3: 99, 4: 99, 5: 99, 6: 99, 7: 99, 8: 99, 9: 99, 10: 99, 11: 99, 12: 99, 13: 99, 14: 99, 15: 99, 16: 99}
chance_card = {1: 39, 2: 0, 3: 30, 4: 24, 5: 11, 6: 5, 7: 42, 8: 42, 9: 41, 10: -3, 11: 99, 12: 99, 13: 99, 14: 99, 15: 99, 16: 99}
jail_turn = {}
player_pos = {}
loc_count = {}
    
comm_chest = shuffle(shuffle(comm_chest))
chance_card = shuffle(shuffle(chance_card))

for i in range(1, num_players+1):
    player_pos[i] = 0
    jail_turn[i] = 0
    
for i in range(0, 40):
    loc_count[i] = 0

for a in range(0, 100):
    for i in range (1, num_players+1): #Last digit of range = # of players - 1
        double_count = 0

        while(double_count < 3):
            roll_1 = rollDice()
            roll_2 = rollDice()
            dice_roll = roll_1+roll_2
            isDouble = False

            if(roll_1 == roll_2):
                double_count += 1
                isDouble = True
            
            if(double_count == 3):
                player_pos[i] = 10
                loc_count[player_pos[i]] += 1
                jail_turn[i] = 3
                isDouble = False
                break
                
            if(jail_turn[i] == 0 or (jail_turn[i] > 0 and isDouble is True)): #if they are not in jail

                jail_turn[i] = 0
                player_pos[i] = (player_pos[i]+dice_roll) % 40

                ''' Use if instead of elif, in case they hit chance, get -3, hit comm chest'''

                if(player_pos[i] == (8 or 22 or 36)): #chance
                    player_pos[i] = newPos(player_pos[i], chance_card[1])
                    chance_card = chance(chance_card)

                if(player_pos[i] == (2 or 17 or 33)): #comm chest
                    player_pos[i] = newPos(player_pos[i], comm_chest[1])
                    comm_chest = community(comm_chest)

                if(player_pos[i] == 30): #don't use elif, because the person can still go to jail from the chance or comm chest card
                    jail_turn[i] = 3
                    player_pos[i] = 10
                
            else:
                jail_turn[i] += -1
                    
            loc_count[player_pos[i]] += 1

            if(not isDouble):
                break

loc_count [8] = loc_count[8] + loc_count[22] + loc_count[36] #combine all the chance
loc_count[22] = -1
loc_count[36] = -1
loc_count [2] = loc_count[2] + loc_count[17] + loc_count[33] #combine all the comm chest
loc_count[17] = -1
loc_count[33] = -1

sum_ = 0
for a in range(1, 40):
    if (loc_count[a] >= 0):
        sum_ += loc_count[a]

ave_count = {}

for a in range(0, 40):
    if (loc_count[a] >= 0):
        ave_count[a] = loc_count[a]/sum_

    else:
        ave_count[a] = -1

for a in range(2, 38):
    
    if (a <= 18):
        sheet.cell(row = a, column = 2).value = loc_count[a-2]
        sheet.cell(row = a, column = 3).value = ave_count[a-2]

    elif (19 <= a <= 22):
        sheet.cell(row = a, column = 2).value = loc_count[a-1]
        sheet.cell(row = a, column = 3).value = ave_count[a-1]

    elif (23 <= a <= 32):
        sheet.cell(row = a, column = 2).value = loc_count[a]
        sheet.cell(row = a, column = 3).value = ave_count[a]

    elif (33 <= a <= 34):
        sheet.cell(row = a, column = 2).value = loc_count[a+1]
        sheet.cell(row = a, column = 3).value = ave_count[a+1]

    elif (35 <= a <= 37):
        sheet.cell(row = a, column = 2).value = loc_count[a+2]
        sheet.cell(row = a, column = 3).value = ave_count[a+2]

    sheet.cell(row = a, column = 3).number_format = FORMAT_PERCENTAGE_00

print('Done!')
wb.save("Monopoly.xlsx")
