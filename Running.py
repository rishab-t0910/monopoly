from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import math
import openpyxl

def changetime(hour, minute, second):

    if(second>=60):
        minute += int(second/60)
        second = second%60

    if(minute>=60):
        hour += int(minute/60)
        minute = minute%60

    if (hour<10):
        hour = '0{}'.format(hour)

    if (minute<10):
        minute = '0{}'.format(minute)

    if (second<10):
        second = '0{}'.format(second)
        
    return '{}:{}:{}'.format(hour, minute, second)

def validDate(day, month):
    trueDay = False
    trueMonth = False
    monthNum = 0
    months = 'janfebmaraprmayjunjulaugsepoctnovdec'

    if(month.isnumeric() and (int(month)<1 or int(month)>12)):
        trueMonth = False

    if(not month.isnumeric()):
        if(month[0:3].lower() in months):
            trueMonth = True
            monthNum = 1+int(months.find(month[0:3].lower(),0)/3)
        else:
            trueMonth = False
        
    if(not day.isnumeric()):
        trueDay = False
        
    if(day.isnumeric()):
        if(int(day)>31 or int(day)<1):
            trueDay = False
        if(monthNum>0):
            if(monthNum == 2): #Feb
                if(1<=int(day)<=28):
                    trueDay = True
                else:
                    trueDay = False
            if(monthNum == (1 or 3 or 5 or 7 or 8 or 10 or 12)):
                if(1<=int(day)<=31):
                    trueDay = True
                else:
                    trueDay = False
            else:
                if(1<=int(day)<=30):
                    trueDay = True
                else:
                    trueDay = False
            
    return trueDay and trueMonth
    
def date(day, month):
    month_name = 'janfebmaraprmayjunjulaugsepoctnovdec'

    if(month.isnumeric()):
        fin_month = month[3*(month-1):3*month]
    else:
        fin_month = month[0:3]

    if(int(day)<10):
        day = '0{}'.format(day)
    
    return '{} {}'.format(day, fin_month)

    
#wb = Workbook()
#wb.save(filename = 'Run Tracker.xlsx')
wb = load_workbook("Run Tracker.xlsx")

sheet = wb.active
confirmation = 0

while(confirmation == 0):
    
    head_1 = sheet.cell(row = 1, column = 1)
    head_1.value = 'Date'
    head_1.font = Font(name = 'Times New Roman', size = 12, bold = True)
    head_1.alignment = Alignment(horizontal="center", vertical="center")

    head_2 = sheet.cell(row = 1, column = 2)
    head_2.value = 'Distance'
    head_2.font = Font(name = 'Times New Roman', size = 12, bold = True)
    head_2.alignment = Alignment(horizontal="center", vertical="center")

    head_3 = sheet.cell(row = 1, column = 3)
    head_3.value = 'Time'
    head_3.font = Font(name = 'Times New Roman', size = 12, bold = True)
    head_3.alignment = Alignment(horizontal="center", vertical="center")

    head_4 = sheet.cell(row = 1, column = 4)
    head_4.value = 'Pace'
    head_4.font = Font(name = 'Times New Roman', size = 12, bold = True)
    head_4.alignment = Alignment(horizontal="center", vertical="center")

    head_5 = sheet.cell(row = 1, column = 5)
    head_5.value = 'Total Time'
    head_5.font = Font(name = 'Times New Roman', size = 12, bold = True)
    head_5.alignment = Alignment(horizontal="center", vertical="center")

    head_6 = sheet.cell(row = 3, column = 7)
    head_6.value = 'Total Distance'
    head_6.font = Font(name = 'Times New Roman', size = 12, bold = True)
    head_6.alignment = Alignment(horizontal="center", vertical="center")

    head_7 = sheet.cell(row = 4, column = 7)
    head_7.value = 'Average Pace'
    head_7.font = Font(name = 'Times New Roman', size = 12, bold = True)
    head_7.alignment = Alignment(horizontal="center", vertical="center")

    sheet.column_dimensions['A'].width = 17
    sheet.column_dimensions['B'].width = 17
    sheet.column_dimensions['C'].width = 17
    sheet.column_dimensions['D'].width = 17
    sheet.column_dimensions['E'].width = 17
    sheet.column_dimensions['G'].width = 17
    sheet.column_dimensions['H'].width = 17

    i = 1

    for cell in sheet['A']:
        if cell.value is None:
            break
        else:
            i = i+1

    date_1 = sheet.cell(row = i, column = 1)
    date_1.font = Font(name = 'Times New Roman', size = 12)
    
    day = input('Day: ')
    month = input('Month: ')
    
    if(not validDate(day, month)):
        print('You have entered an invalid date. Thank you')
        break

    else:
        date_1.value = date(day, month).title()
        
    date_1.alignment = Alignment(horizontal="center", vertical="center") 

    dist_1 = sheet.cell(row = i, column = 2)
    dist_1.value = float(input('Distance (km): '))
    meters = dist_1.value*1000 #convert to meters
    dist_1.font = Font(name = 'Times New Roman', size = 12)
    dist_1.alignment = Alignment(horizontal="center", vertical="center")
    dist_1.number_format = '00.00'

    time_1 = sheet.cell(row = i, column = 3)

    cum_time = 0 #cummalitive time for each run

    hour = int(input('Hours: '))
    cum_time+=3600*hour

    minutes = int(input('Minutes: '))
    cum_time+=60*minutes

    seconds = int(input('Seconds: '))
    cum_time+=seconds
    
    time_1.number_format = '%H:%M:%S'
    time_1.value = changetime(hour, minutes, seconds) #if 90 minute input, return is 01:30:00
    time_1.font = Font(name = 'Times New Roman', size = 12)
    time_1.alignment = Alignment(horizontal="center", vertical="center")

    total_time = sheet.cell(row = i, column = 5)
    prev_time = sheet.cell(row = i-1, column = 5)
    gross_time = 0 #total time ever

    if(i==2):
        total_time.value = time_1.value
        total_time.number_format = '%H:%M:%S'
        gross_time = int(3600*hour)+int(60*minutes)+int(seconds)
    else:
        prev_time = sheet.cell(row = i-1, column = 5)
        prev_time.number_format = '%H:%M:%S'
        hh = int(prev_time.value[0:2])+int(time_1.value[0:2])
        mm = int(prev_time.value[3:5])+int(time_1.value[3:5])
        ss = int(prev_time.value[6:8])+int(time_1.value[6:8])
        gross_time = int(3600*hh)+int(60*mm)+int(ss) 
        total_time.number_format = '%H:%M:%S'

        if(ss>=60):
            mm+=int(ss/60)
            ss = ss%60
        if(mm>=60):
            hh+=int(mm/60)
            mm = mm%60

        if (hh<10):
            hh = '0{}'.format(hh)
        if (mm<10):
            mm = '0{}'.format(mm)
        if (ss<10):
            ss = '0{}'.format(ss)
        
        total_time.value = '{}:{}:{}'.format(hh, mm, ss)

    #gross_time is in seconds now

    total_time.font = Font(name = 'Times New Roman', size = 12)
    total_time.alignment = Alignment(horizontal="center", vertical="center")

    single_pace = float((cum_time*1000)/(meters*60))
    min_pace = int(math.floor(single_pace))
    sec_pace = int(60*float(float(single_pace) - float(min_pace)))
    if(min_pace<10):
        min_pace = '0{}'.format(min_pace)
    if(sec_pace<10):
        sec_pace = '0{}'.format(sec_pace)

    pace_1 = sheet.cell(row = i, column = 4)
    pace_1.value = '{}:{}'.format(min_pace, sec_pace)
    pace_1.font = Font(name = 'Times New Roman', size = 12)
    pace_1.alignment = Alignment(horizontal="center", vertical="center")

    sheet['H3'] = '= SUM(B:B)'
    sheet['H3'].font = Font(name = 'Times New Roman', size = 12)
    sheet['H3'].alignment = Alignment(horizontal="center", vertical="center")
    sheet['H3'].number_format = '00.00'

    gross_dist = 0 #total distance ever
    i = 1
    for cell in sheet['B']:
        if i == 1: #skip the heading
            i = i+1
        elif cell.value is None: #end 
            break
        else:
            gross_dist+= 1000*int(cell.value) #convert to meters

    #gross_dist in meter, gross_time in seconds

    gross_time = gross_time*1000
    gross_dist = gross_dist*60
    ave_pace = float(gross_time)/float(gross_dist)
    ave_min_pace = int(math.floor(ave_pace))
    ave_sec_pace = int(60*float(ave_pace-ave_min_pace))

    if(ave_min_pace<10):
        ave_min_pace = '0{}'.format(ave_min_pace)

    if(ave_sec_pace<10):
        ave_sec_pace = '0{}'.format(ave_sec_pace)

    average_pace = sheet.cell(row = 4, column = 8)
    average_pace.value = '{}:{}'.format(ave_min_pace, ave_sec_pace)
    average_pace.font = Font(name = 'Times New Roman', size = 12)
    average_pace.alignment = Alignment(horizontal="center", vertical="center")

    print('\n')
    print("\u0332".join("Summary"), end = "\n\n")
    print(' Date: ', str(date_1.value)+'\n', 'Duration: ', str(time_1.value)+'\n', 'Distance: ', str(dist_1.value)+'\n', 'Pace: ', str(pace_1.value)+'\n')
    confirmation = int(input('Re-input: 0\nConfirm: 1\nExit: 2\n\n'))

    if (confirmation == 0 or confirmation == 2):
        
        i = 1

        for cell in sheet['A']:
            if cell.value is None:
                break
            else:
                i = i+1
        if(confirmation == 0):
            sheet['H4'].value = None #average pace = None only if re-input
            
        i = i-1
        for col in range(1, 6): #delete value if exit or re-input
           cell_obj = sheet.cell(row = i, column = col)
           cell_obj.value = None
           
    if(confirmation == 2):
        
        print('Thank you')
        break

wb.save("Run Tracker.xlsx")
