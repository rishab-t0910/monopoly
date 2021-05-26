from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import math
import openpyxl

def floatCheck(value):

    try:
        value = float(value)
        return 1
    
    except ValueError:
        return 0

def intCheck(value):

    try:
        value = int(value)
        return 1

    except ValueError:
        return 0
    
def changetime(hour, minute, second):

    second = int(second)
    minute += int(second/60)
    second = second % 60
    hour += int(minute/60)
    minute = minute % 60
    
    if (hour < 10):
        hour = '0{}'.format(hour)

    if (minute < 10):
        minute = '0{}'.format(minute)

    if (second < 10):
        second = '0{}'.format(second)

    return '{}:{}:{}'.format(hour, minute, second)

def validDate(day, month):
    
    trueMonth = False
    trueDay = False
    
    monthName = 'janfebmaraprmayjunjulaugsepoctnovdec'
    fin_month = ''

    month_num = 0

    if (intCheck(month)):

        if (not (1 <= int(month) <= 12)):
            trueMonth = False
            
        else:
            fin_month = monthName[3*(int(month)-1):3*int(month)]
            month_num = int(month)
            trueMonth = True
            
    else:
        month = str(month[0:3])
        
        if (monthName.find(month) >= 0 and monthName.find(month) % 3 == 0):
            month_num = int(monthName.find(month)/3) + 1
            fin_month = month
            trueMonth = True
            
        else:
            trueMonth = False

    if (intCheck(day)):

        if (not (1 <= int(day) <= 31)):
            trueDay = False
            
        else: #month by month case
            day = int(day)
            
            if (month_num == 2):
                trueDay = (1 <= day <= 29)

            elif (month_num == 1 or month_num == 3 or month_num == 5 or month_num == 7 or month_num == 8 or month_num == 10 or month_num == 12):
                trueDay = (1 <= day <= 31)
                
            else:
                trueDay = (1 <= day <= 30)

    else:
        trueDay = False

    return (trueDay and trueMonth)

def date(day, month):
    
    month_name = 'janfebmaraprmayjunjulaugsepoctnovdec'
    month = str(month)

    if (month.isnumeric() and 1 <= int(month) <= 12):
        fin_month = month_name[3*(int(month)-1):3*int(month)]

    else:
        fin_month = month_name[0:3]

    if (int(day) < 10):
        day = '0{}'.format(day)

    return '{} {}'.format(day.title(), fin_month.title())

#wb = Workbook()
#wb.save(filename = 'Run Tracker.xlsx')
wb = load_workbook("Run Tracker.xlsx")

sheet = wb.active

print('Welcome to the Run Tracker! What would you like to do first?\n\n')

choice = str(input('0: Input\n1: Find\n2: Show all runs\n3: Exit\n'))

count = 1

while (not(intCheck(choice) and 0 <= int(choice) <= 3) and count < 4): #check valid user input
    
    print('You have entered an invalid choice.')
    count += 1
    choice = input('0: Input\n1: Find\n2: Show all runs\n3: Exit\n')

if (not(intCheck(choice) and 0 <= int(choice) <= 3)):
    print('Goodbye.')

else:
    print('')
    choice = int(choice)
    
    if (choice == 1):

        print('What would you like to find?\n')
        data_point = str(input('0: Highest Run Distance\n1: Highest Run Time\n2: Run Stats on a specific date\n3: Total Distance\n4: Total Time\n5: Total Average Pace\n6: Lowest Pace\n7: Runs within a range of distance\n8: Runs within a range of dates\n9: Runs within a range of times\n10: Exit\n'))

        count = 1

        while (not(intCheck(choice) and 0 <= int(data_point) <= 10) and count < 4): #check valid user input
            print('You have entered an invalid choice.')
            count += 1
            data_point = str(input('0: Highest Run Distance\n1: Highest Run Time\n2: Run Stats on a specific date\n3: Total Distance\n4: Total Time\n5: Total Average Pace\n6: Lowest Pace\n7: Runs within a range of distance\n8: Runs within a range of dates\n9: Runs within a range of times\n10: Exit\n'))

        if (not(intCheck(choice) and 0 <= int(data_point) <= 10)):
            print('Goodbye.')

        else:
            print('')
            data_point = int(data_point)

            if (data_point == 0): #highest run dist, second column = B

                max_val = -1
                i = 1
                
                for cell in sheet['B']:
                    if (cell.value is None):
                        break
                    
                    else:
                        if (not (floatCheck(cell.value) and intCheck(cell.value)):
                            i += 1

                        else:
                            if (float(cell.value) > max_val):
                                max_val = cell.value
                                
                            i += 1

                row_count = 1
                iter_1 = 0
                
                date_dict = {}
                time_dict = {}
                pace_dict = {}

                for cell in sheet['B']:
                    if (cell.value is None):
                        break

                    if (row_count == 1):
                        pass

                    else:
                        if (float(cell.value) == max_val):
                            date_dict[iter_1] = sheet.cell(row = row_count, column = 1).value
                            time_dict[iter_1] = sheet.cell(row = row_count, column = 3).value
                            pace_dict[iter_1] = sheet.cell(row = row_count, column = 4).value
                            iter_1 += 1

                    row_count += 1

                if (max_val == -1):
                    print('There are no entries.')

                else:
                    print('All runs with the higest distance of {}km are:'.format(max_val))
                    print("\u0332".join("Date"), '\t','\t',"\u0332".join("Time"), '\t\t',"\u0332".join("Pace"))

                    for i in range (0, len(pace_dict)):
                        print(date_dict[i], '\t\t', time_dict[i], '\t', pace_dict[i])

            elif (data_point == 1): #high run time, third column = C

                max_val = '00:00:00'
                
                i = 0

                for cell in sheet['C']:
                    if(max_val is None or cell.value is None):
                        break
                    
                    else:
                        if (i == 0):
                            i += 1

                        else:
                            if (int(cell.value[0:2]) > int(max_val[0:2])): #checks hours
                               max_val = cell.value

                            elif (int(cell.value[0:2]) == int(max_val[0:2])): #if hours are same
                                if (int(cell.value[3:5]) > int(max_val[3:5])): #checks min
                                    max_val = cell.value

                                elif (int(cell.value[3:5]) == int(max_val[3:5])): #if min are same
                                    if (int(cell.value[6:8]) > int(max_val[6:8])): #checks sec
                                        max_val = cell.value
                            i += 1

                iter_1 = 0
                row_ = 1
                
                date_dict = {}
                dist_dict = {}
                pace_dict = {}

                for cell in sheet['C']:

                    if (cell.value is None):
                        break

                    if (row_ == 1):
                        pass

                    else:
                        if (int(cell.value[0:2]) == int(max_val[0:2]) and int(cell.value[3:5]) == int(max_val[3:5]) and int(cell.value[6:8]) == int(max_val[6:8])):
                            date_dict[iter_1] = sheet.cell(row = row_, column = 1).value
                            dist_dict[iter_1] = sheet.cell(row = row_, column = 2).value
                            pace_dict[iter_1] = sheet.cell(row = row_, column = 4).value
                            iter_1 +=  1

                    row_ += 1



                if(max_val is None):
                    print('There are no entries.')

                else:
                    print('All runs with the higest run time of {} are:'.format(max_val))
                    print("\u0332".join("Date"), '\t','\t',"\u0332".join("Distance"), '\t',"\u0332".join("Pace"))

                    for i in range (0, len(pace_dict)):
                        print(date_dict[i], '\t\t', str(dist_dict[i])+'km', '\t\t', pace_dict[i])

            elif (data_point == 2): #run stat on date, first column = A
                day = input('Day: ')
                month = input('Month: ')

                count = 1

                while (not validDate(day, month) and count < 4):
                    print('Invalid Date.')
                    day = input('Day: ')
                    month = input('Month: ')
                    count += 1

                if (not validDate(day, month)):
                    print('Goodbye.')

                else:
                    date_check = date(day, month)
                    
                    i = 1
                    correct = 0
                    iter_1 = 0

                    dist_dict = {}
                    time_dict = {}
                    pace_dict = {}

                    for cell in sheet['A']:
                        if (cell.value is None):
                            break

                        else:
                            if (cell.value == date_check):
                                correct = 1
                                dist_dict[iter_1] = sheet.cell(row = i, column = 2).value
                                time_dict[iter_1] = sheet.cell(row = i, column = 3).value
                                pace_dict[iter_1] = sheet.cell(row = i, column = 4).value
                                iter_1 += 1

                            i += 1

                    if (correct == 0):
                        print('This date does not exist.')

                    else:
                        print('All runs on {} are:'.format(date_check))
                        print("\u0332".join("Distance"), '\t',"\u0332".join("Time"), '\t\t',"\u0332".join("Pace"))

                    for i in range (0, len(pace_dict)):
                        print(str(dist_dict[i])+'km', '\t\t', time_dict[i], '\t', pace_dict[i])

            elif (data_point == 3): #total dist, row = 3, column = 8, H                
                total_dist = 0
                
                for cell in sheet['B']:
                    
                    if (cell.value is not None):
                        
                        if (floatCheck(cell.value) or intCheck(cell.value)):
                            total_dist += float(cell.value)

                print('Total Distance:', str(total_dist)+'km.\n')

            elif (data_point == 4): #total time, last row on column = 5, E
                i = 0

                for cell in sheet['E']:
                    if (cell.value is None):
                        break

                    i += 1

                total_time = sheet.cell(row = i, column = 5).value

                hour = total_time[0:2]
                minute = total_time[3:5]
                sec = total_time[6:8]

                valid_hour = 0
                valid_min = 0
                valid_sec = 0

                if (intCheck(hour)):
                    valid_hour = 1
                    if (int(hour) == 1):
                        hour = str(hour)+' hour'

                    else:
                        hour = str(hour)+' hours'

                if (intCheck(minute)):
                    valid_min = 1
                    if (int(minute) == 1):
                        minute = str(minute)+' minute'

                    else:
                        minute = str(minute)+' minutes'

                if (intCheck(sec)):
                    valid_sec = 1
                    if (int(sec) == 1):
                        sec = str(sec)+' second'

                    else:
                        sec = str(sec)+' seconds'

                if (valid_sec and valid_min and valid_hour): #check if there is a total time value
                    print('Total Time: {}, {}, {}.\n'.format(hour, minute, sec))

                else:
                    print('There are no entries.')

            elif (data_point == 5): #total ave pace, row = 4, column = 8, H
                ave_pace = sheet.cell(row = 4, column = 8).value

                if (ave_pace is not None):
                    minute = ave_pace[0:2]
                    sec = ave_pace[3:5]

                    if (int(minute) == 1):
                        minute = str(minute)+' minute'

                    else:
                        minute = str(minute)+' minutes'

                    if (int(sec) == 1):
                        sec = str(sec)+' second'

                    else:
                        sec = str(sec)+' seconds'

                    print('Total Average pace (minutes per kilometer): {}, {}.\n'.format(minute, sec))

                else:
                    print('There are no entries.')

            elif (data_point == 6): #best pace, last row, column = 4, D
                max_val = '99:99'
                i = 0
                
                for cell in sheet['D']:
                    if (max_val is None or cell.value is None):
                        break
                    
                    else:
                        if (i == 0):
                            i += 1

                        else:
                            if (int(cell.value[0:2]) < int(max_val[0:2])): #checks min
                               max_val = cell.value

                            elif (int(cell.value[0:2]) == int(max_val[0:2])): #if min are same
                                if (int(cell.value[3:5]) < int(max_val[3:5])): #checks sec
                                    max_val = cell.value
                            i += 1

                iter_1 = 0
                row_ = 1

                date_dict = {}
                dist_dict = {}
                time_dict = {}

                for cell in sheet['D']:

                    if (cell.value is None):
                        break

                    if (row_ == 1):
                        pass

                    else:
                        if (int(cell.value[0:2]) == int(max_val[0:2]) and int(cell.value[3:5]) == int(max_val[3:5])):
                            date_dict[iter_1] = sheet.cell(row = row_, column = 1).value
                            dist_dict[iter_1] = sheet.cell(row = row_, column = 2).value
                            time_dict[iter_1] = sheet.cell(row = row_, column = 3).value
                            iter_1 +=  1

                    row_ += 1



                if (max_val is None):
                    print('There are no entries.')

                else:
                    print('All runs with the best pace of {} are:'.format(max_val))
                    print("\u0332".join("Date"), '\t','\t',"\u0332".join("Distance"), '\t',"\u0332".join("Time"))

                    for i in range (0, len(time_dict)):
                        print(date_dict[i], '\t\t', str(dist_dict[i])+'km', '\t\t', time_dict[i])

            elif (data_point == 7): #distance range
                lower_bound = input('Enter a lower distance bound: ')
                
                count = 1
                
                while (not intCheck(lower_bound) and count < 4):
                    
                    print('Invalid Lower Bound.')
                    count += 1
                    lower_bound = input('Enter a lower distance bound: ')

                if (not intCheck(lower_bound)):
                    print('Goodbye.')
                    
                else:
                    upper_bound = input('Enter a upper distance bound: ')
                    
                    count = 1
                    
                    while(not intCheck(upper_bound) and count < 4):
                        print('Invalid upper Bound.')
                        count += 1
                        upper_bound = input('Enter a upper distance bound: ')
                        
                    if (not intCheck(upper_bound)):
                        print('Goodbye.')
                        
                    else:
                        iter_1 = 0
                        rows = 1
                        
                        date_dict = {}
                        dist_dict = {}
                        time_dict = {}
                        pace_dict = {}

                        for cell in sheet['B']:
                            if (cell.value is None):
                                break

                            if (rows == 1):
                                rows += 1

                            else:
                                if (float(lower_bound) <= float(cell.value) <= float(upper_bound)):
                                    date_dict[iter_1] = sheet.cell(row = rows, column = 1).value
                                    dist_dict[iter_1] = sheet.cell(row = rows, column = 2).value
                                    time_dict[iter_1] = sheet.cell(row = rows, column = 3).value
                                    pace_dict[iter_1] = sheet.cell(row = rows, column = 4).value
                                    iter_1 += 1

                                rows += 1

                    if (len(date_dict) == 0):
                        print('There are no entries between {}km and {}km.'.format(lower_bound, upper_bound))

                    else:
                        print('\nAll runs between {}km and {}km are:'.format(lower_bound, upper_bound))
                        print("\u0332".join("Date"), '\t', "\u0332".join("Distance"),'\t',"\u0332".join("Time"), '\t\t',"\u0332".join("Pace"))

                        for i in range (0, len(date_dict)):
                            print(date_dict[i], '\t' , dist_dict[i],'km\t\t', time_dict[i], '\t', pace_dict[i])


            elif (data_point == 8): #date range
                print('Enter a valid start date')
                day = input('Day: ')
                month = input('Month: ')
                print('\n')
                
                count = 1

                while (not validDate(day, month) and count < 4):
                    
                    print('Enter a valid start date')
                    count += 1
                    day = input('Day: ')
                    month = input('Month: ')
                    print('\n')

                if (not validDate(day, month)):
                    print('Goodbye.')
                    
                else:
                    start_date = date(day, month)

                    print('Enter a valid end date')
                    day = input('Day: ')
                    month = input('Month: ')
                    print('\n')
                    count = 1

                    while (not validDate(day, month) and count < 4):
                        
                        print('Enter a valid end date')
                        count += 1
                        day = input('Day: ')
                        month = input('Month: ')
                        print('\n')

                    if (not validDate(day, month)):
                        print('Goodbye.')
                        
                    else:
                        end_date = date(day, month) #Creates date
                        months = 'janfebmaraprmayjunjulaugsepoctnovdec'

                        iter_1 = 0
                        rows = 1
                        to_add = 0
                        start_day = int(start_date[0:2])
                        start_month = int(round(months.find(start_date[3:6].lower())/4))+1
                        end_day = int(end_date[0:2])
                        end_month = int(round(months.find(end_date[3:6].lower())/4))+1
                        
                        date_dict = {}
                        dist_dict = {}
                        time_dict = {}
                        pace_dict = {}

                        for cell in sheet['A']:
                            to_add = 0

                            if (cell.value is None):
                                break

                            if (rows == 1):
                                rows += 1

                            else:

                                curr_month = int(round(months.find(cell.value[3:6].lower())/4))+1
                                curr_date = int(cell.value[0:2])

                                if (start_month < curr_month < end_month):
                                    to_add = 1

                                elif (start_month == curr_month == end_month):
                                    to_add = (start_day <= curr_date <= end_day)

                                elif (start_month == curr_month):
                                    to_add = (curr_date >= start_day)
                                    
                                elif (end_month == curr_month):
                                    to_add = (curr_date <= end_day):

                                if (to_add == 1):
                                    date_dict[iter_1] = sheet.cell(row = rows, column = 1).value
                                    dist_dict[iter_1] = sheet.cell(row = rows, column = 2).value
                                    time_dict[iter_1] = sheet.cell(row = rows, column = 3).value
                                    pace_dict[iter_1] = sheet.cell(row = rows, column = 4).value
                                    iter_1 += 1

                                rows += 1

                        if(len(date_dict) == 0):
                            print('There are no entries between {} and {}.'.format(start_date, end_date))

                        else:
                            print('All runs between {} and {} are:'.format(start_date, end_date))
                            print("\u0332".join("Date"), '\t', "\u0332".join("Distance"),'\t',"\u0332".join("Time"), '\t\t',"\u0332".join("Pace"))

                            for i in range (0, len(date_dict)):
                                  print(date_dict[i], '\t' , dist_dict[i],'km\t\t', time_dict[i], '\t', pace_dict[i])

            elif(data_point == 9): #time range
                hour = input('Enter a start hour: ')
                minute = input('Enter a start minute: ')
                second = input('Enter a start second: ')
                count = 1

                while (not (intCheck(hour) and intCheck(minute) and intCheck(second)) and count < 4):
                    print('Enter a valid start time')
                    count += 1
                    hour = input('Enter a start hour: ')
                    minute = input('Enter a start minute: ')
                    second = input('Enter a start second: ')


                if (not (intCheck(hour) and intCheck(minute) and intCheck(second))):
                    print('Goodbye.')
                    
                else:
                    start_time = changetime(hour, minute, second) #Even if invalid time is given, we adjust. Eg: 0, 90, 0 input ==> 1, 30, 0 output

                    start_hour = int(start_time[0:2])
                    start_min = int(start_time[3:5])
                    start_sec = int(start_time[6:8])

                    hour = input('\nEnter an end hour: ')
                    minute = input('Enter an end minute: ')
                    second = input('Enter an end second: ')
                    count = 1

                    while (not (intCheck(hour) and intCheck(minute) and intCheck(second)) and count < 4):
                        print('Enter a valid end time')
                        count += 1
                        hour = input('Enter an end hour: ')
                        minute = input('Enter an end minute: ')
                        second = input('Enter an end second: ')

                    if (not (intCheck(hour) and intCheck(minute) and intCheck(second))):
                        print('Goodbye.')

                    else:
                        end_time = changetime(hour, minute, second)
                        end_hour = int(end_time[0:2])
                        end_min = int(end_time[3:5])
                        end_sec = int(end_time[6:8])
                        iter_1 = 0
                        rows = 1
                        to_add = 0

                        date_dict = {}
                        dist_dict = {}
                        time_dict = {}
                        pace_dict = {}
                        
                        for cell in sheet['C']:
                            to_add = 0
                            if (cell.value is None):
                                break

                            if (rows == 1):
                                rows += 1

                            else:
                                cell_hour = int(cell.value[0:2])
                                cell_min = int(cell.value[3:5])
                                cell_sec = int(cell.value[6:8])

                                if (cell_hour > end_hour or cell_hour < start_hour):
                                    to_add = 0

                                elif (cell_hour == start_hour and cell_hour == end_hour):
                                    if (cell_min > start_min and cell_min < end_min):
                                        to_add = 1 

                                    elif (cell_min == start_min and cell_min == end_min):
                                        to_add = (start_sec <= cell_sec <= end_sec)
                                        
                                    else:
                                        to_add = 0

                                elif (cell_hour == start_hour):
                                    if (cell_min < start_min):
                                        to_add = 0

                                    elif (cell_min > start_min):
                                        to_add = 1

                                    else:
                                        to_add = (cell_sec >= start_sec)

                                elif (cell_hour == end_hour): 
                                    if (cell_min > end_min):
                                        to_add = 0

                                    elif (cell_min < end_min):
                                        to_add = 1

                                    else:
                                        to_add = (cell_sec <= end_sec)

                                elif (start_hour < cell_hour < end_hour):
                                    to_add = 1

                                if (to_add == 1):
                                    date_dict[iter_1] = sheet.cell(row = rows, column = 1).value
                                    dist_dict[iter_1] = sheet.cell(row = rows, column = 2).value
                                    time_dict[iter_1] = sheet.cell(row = rows, column = 3).value
                                    pace_dict[iter_1] = sheet.cell(row = rows, column = 4).value
                                    iter_1 += 1

                                rows += 1

                        if(len(date_dict) == 0):
                            print('\nThere are no entries between {} and {}.'.format(start_time, end_time))

                        else:
                            print('\nAll runs between {} and {} are:\n'.format(start_time, end_time))
                            print("\u0332".join("Date"), '\t', "\u0332".join("Distance"),'\t',"\u0332".join("Time"), '\t\t',"\u0332".join("Pace"))

                            for i in range (0, len(date_dict)):
                                  print(date_dict[i], '\t' , dist_dict[i],'km\t\t', time_dict[i], '\t', pace_dict[i])

            else: #exit
                print('Goodbye.')

    elif (choice == 0):
        confirmation = 0

        while(confirmation == 0):

            head_1 = sheet.cell(row = 1, column = 1)
            head_1.value = 'Date'
            head_1.font = Font(name = 'Times New Roman', size = 12, bold = True)
            head_1.alignment = Alignment(horizontal="center", vertical="center")

            head_2 = sheet.cell(row = 1, column = 2)
            head_2.value = 'Distance'
            head_2.font = Font(name = 'Times New Roman', size = 12, bold = True)
            head_2.alignment = Alignment(horizontal="center", vertical="center")

            head_3 = sheet.cell(row = 1, column = 3)
            head_3.value = 'Time'
            head_3.font = Font(name = 'Times New Roman', size = 12, bold = True)
            head_3.alignment = Alignment(horizontal="center", vertical="center")

            head_4 = sheet.cell(row = 1, column = 4)
            head_4.value = 'Pace'
            head_4.font = Font(name = 'Times New Roman', size = 12, bold = True)
            head_4.alignment = Alignment(horizontal="center", vertical="center")

            head_5 = sheet.cell(row = 1, column = 5)
            head_5.value = 'Total Time'
            head_5.font = Font(name = 'Times New Roman', size = 12, bold = True)
            head_5.alignment = Alignment(horizontal="center", vertical="center")

            head_6 = sheet.cell(row = 3, column = 7)
            head_6.value = 'Total Distance'
            head_6.font = Font(name = 'Times New Roman', size = 12, bold = True)
            head_6.alignment = Alignment(horizontal="center", vertical="center")

            head_7 = sheet.cell(row = 4, column = 7)
            head_7.value = 'Average Pace'
            head_7.font = Font(name = 'Times New Roman', size = 12, bold = True)
            head_7.alignment = Alignment(horizontal="center", vertical="center")

            sheet.column_dimensions['A'].width = 17
            sheet.column_dimensions['B'].width = 17
            sheet.column_dimensions['C'].width = 17
            sheet.column_dimensions['D'].width = 17
            sheet.column_dimensions['E'].width = 17
            sheet.column_dimensions['G'].width = 17
            sheet.column_dimensions['H'].width = 17

            i = 1

            for cell in sheet['A']:
                if cell.value is None:
                    break

                else:
                    i += 1

            date_1 = sheet.cell(row = i, column = 1)
            date_1.font = Font(name = 'Times New Roman', size = 12)

            day = input('Day: ')
            month = input('Month: ')
            count = 1

            while (not validDate(day, month) and count < 4):
                
                print('Enter a valid date.')
                count += 1
                day = input('Day: ')
                month = input('Month: ')

            if (not validDate(day, month)):
                print('Goodbye.')
                break

            else:
                date_1.value = date(day, month).title()
                date_1.alignment = Alignment(horizontal="center", vertical="center")
                dist_1 = sheet.cell(row = i, column = 2)

                distance = input('Distance (kilometer): ')

                count = 1

                while (not (floatCheck(distance) and float(distance) > 0) and count < 4):
                    print('Invalid distance.')
                    count += 1
                    distance = input('Distance (kilometer): ')

                if (not (floatCheck(distance) and float(distance) > 0)):
                    sheet.cell(row = i, column = 1).value = None
                    print('Goodbye.')
                    break

                else:
                    dist_1.value = float(distance)
                    
                meters = dist_1.value*1000 #convert to meters
                
                dist_1.font = Font(name = 'Times New Roman', size = 12)
                dist_1.alignment = Alignment(horizontal="center", vertical="center")
                dist_1.number_format = '00.00'

                time_1 = sheet.cell(row = i, column = 3)

                cum_time = 0 #cummalitive time for each run

                hour = input('Hours: ')
                minutes = input('Minutes: ')
                seconds = input('Seconds: ')
                
                count = 1
                
                while (not (intCheck(hour) and intCheck(minute) and intCheck(second)) or (hour == 0 and minutes == 0 and seconds == 0) and count < 4):
                    print('You have entered an invalid time.')
                    count += 1
                    hour = input('Hours: ')
                    minutes = input('Minutes: ')
                    seconds = input('Seconds: ')
                    
                if (not (intCheck(hour) and intCheck(minute) and intCheck(second)) or (hour == 0 and minutes == 0 and seconds == 0)):
                    sheet.cell(row = i, column = 1).value = None
                    sheet.cell(row = i, column = 2).value = None
                    print('Goodbye.')
                    break
                
                else:
                    hour = int(hour)
                    cum_time += 3600*hour
                    minutes = int(minutes)
                    cum_time += 60*minutes
                    seconds = int(seconds)
                    cum_time += seconds
                    
                time_1.number_format = '%H:%M:%S'
                time_1.value = changetime(hour, minutes, seconds) #if 90 minute input, return is 01:30:00
                time_1.font = Font(name = 'Times New Roman', size = 12)
                time_1.alignment = Alignment(horizontal="center", vertical="center")

                total_time = sheet.cell(row = i, column = 5)
                prev_time = sheet.cell(row = i-1, column = 5)
                gross_time = 0 #total time ever

                if (i == 2):
                    total_time.value = time_1.value
                    total_time.number_format = '%H:%M:%S'
                    gross_time = int(3600*hour)+int(60*minutes)+int(seconds)

                else:
                    prev_time = sheet.cell(row = i-1, column = 5)
                    prev_time.number_format = '%H:%M:%S'
                    hh = int(prev_time.value[0:2])+int(time_1.value[0:2])
                    mm = int(prev_time.value[3:5])+int(time_1.value[3:5])
                    ss = int(prev_time.value[6:8])+int(time_1.value[6:8])
                    gross_time = int(3600*hh) + int(60*mm) + int(ss)
                    total_time.number_format = '%H:%M:%S'

                    mm += int(ss/60)
                    ss = ss % 60
                    hh += int(hh/60)
                    mm = mm % 60

                    if (hh < 10):
                        hh = '0{}'.format(hh)

                    if (mm < 10):
                        mm = '0{}'.format(mm)

                    if (ss < 10):
                        ss = '0{}'.format(ss)

                    total_time.value = '{}:{}:{}'.format(hh, mm, ss)

                #gross_time is in seconds now

                total_time.font = Font(name = 'Times New Roman', size = 12)
                total_time.alignment = Alignment(horizontal="center", vertical="center")

                single_pace = float((cum_time*1000)/(meters*60))
                min_pace = int(math.floor(single_pace))
                sec_pace = int(60*float(float(single_pace) - float(min_pace)))
                
                if (min_pace < 10):
                    min_pace = '0{}'.format(min_pace)

                if (sec_pace < 10):
                    sec_pace = '0{}'.format(sec_pace)

                pace_1 = sheet.cell(row = i, column = 4)
                pace_1.value = '{}:{}'.format(min_pace, sec_pace)
                pace_1.font = Font(name = 'Times New Roman', size = 12)
                pace_1.alignment = Alignment(horizontal="center", vertical="center")

                sheet['H3'] = '= SUM(B:B)'
                sheet['H3'].font = Font(name = 'Times New Roman', size = 12)
                sheet['H3'].alignment = Alignment(horizontal="center", vertical="center")
                sheet['H3'].number_format = '00.00'

                gross_dist = 0 #total distance ever
                i = 1
                          
                for cell in sheet['B']:
                    if (i == 1): #skip the heading
                        i += 1

                    elif cell.value is None: #end
                        break

                    else:
                        gross_dist += 1000*int(cell.value) #convert to meters

                #gross_dist in meter, gross_time in seconds

                gross_time = gross_time*1000
                gross_dist = gross_dist*60
                ave_pace = float(gross_time)/float(gross_dist)
                ave_min_pace = int(math.floor(ave_pace))
                ave_sec_pace = int(60*float(ave_pace-ave_min_pace))

                if (ave_min_pace < 10):
                    ave_min_pace = '0{}'.format(ave_min_pace)

                if (ave_sec_pace < 10):
                    ave_sec_pace = '0{}'.format(ave_sec_pace)

                average_pace = sheet.cell(row = 4, column = 8)
                average_pace.value = '{}:{}'.format(ave_min_pace, ave_sec_pace)
                average_pace.font = Font(name = 'Times New Roman', size = 12)
                average_pace.alignment = Alignment(horizontal="center", vertical="center")

                print('\n',"\u0332".join("Summary"), end = "\n\n")
                print(' Date: ', str(date_1.value)+'\n', 'Duration: ', str(time_1.value)+'\n', 'Distance: ', str(dist_1.value)+'\n', 'Pace: ', str(pace_1.value)+'\n')
                confirmation = input('0: Re-input\n1: Confirm\n2: Exit\n\n')

                count = 1

                while (not (intCheck(confirmation) and 0 <= int(confirmation) <= 2) and count < 4): #check valid user input
                    print('You have entered an invalid choice.')
                    count += 1
                    confirmation = input('0: Re-input\n1: Confirm\n2: Exit\n\n')

                if (not (intCheck(confirmation) and 0 <= int(confirmation) <= 2)):
                    confirmation = 2

                confirmation = int(confirmation)

                if (confirmation % 2 == 0):
                    i = 1

                    for cell in sheet['A']:
                        if (cell.value is None):
                            break

                        else:
                            i = i+1


                    if (confirmation == 0):
                        sheet['H4'].value = None #average pace = None only if re-input

                    i = i-1

                    for col in range(1, 6): #delete value if exit or re-input
                       cell_obj = sheet.cell(row = i, column = col)
                       cell_obj.value = None

                    if (confirmation == 2):
                        print('Thank you.')
                        break

                else:
                    print('Congrats on your run!')
                    
    elif (choice == 2): #print all runs
        i = 1

        if (sheet['A2'].value is None):
            print('There are no entries.')

        else:
            print("\u0332".join("Date"), '\t', "\u0332".join("Distance"),'\t',"\u0332".join("Time"), '\t\t',"\u0332".join("Pace"))

            for cell in sheet['A']:
                if (cell.value is None):
                    break

                else:
                    if (i == 1):
                        i += 1

                    else:
                        date_val = sheet.cell(row = i, column = 1).value
                        dist_val = sheet.cell(row = i, column = 2).value
                        time_val = sheet.cell(row = i, column = 3).value
                        pace_val = sheet.cell(row = i, column = 4).value
                        i += 1
                        print(date_val, '\t' , dist_val,'km\t\t', time_val, '\t', pace_val)
        
    else:
        print('Goodbye.')

wb.save("Run Tracker.xlsx")
